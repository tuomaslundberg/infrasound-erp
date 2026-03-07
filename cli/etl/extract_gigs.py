#!/usr/bin/env python3
"""
cli/etl/extract_gigs.py

Extracts and normalises gig, customer, contact, and venue data from all
old-files/info/gigs-YYYY.xlsx (and archive/gigs-YYYY.xlsx) files, plus
old-files/gig-invoicing.xlsx.

Outputs
-------
  db/seeds/legacy_gigs.sql
      Idempotent INSERT statements.  IDs 1–9999 are reserved for this import;
      the file clears and re-inserts that range on every run.  AUTO_INCREMENT
      is reset to 10000 so ERP-created records never collide.

  db/seeds/legacy_gigs_dup_candidates.txt  (if any found)
      Near-duplicate customer name pairs for manual review in the ERP.

Re-running the script as Excel files are updated is safe and intentional.

Usage
-----
  python cli/etl/extract_gigs.py           # write files
  python cli/etl/extract_gigs.py --dry-run # print SQL to stdout
  python cli/etl/extract_gigs.py --stats   # print entity counts only

Prerequisites
-------------
  pip install openpyxl
  DB migration 001_expand_channel_enum.sql must be applied before loading the SQL.
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent.parent

GIGS_GLOBS = [
    str(REPO_ROOT / "old-files/info/gigs-*.xlsx"),
    str(REPO_ROOT / "old-files/info/archive/gigs-*.xlsx"),
]
INVOICING_PATH = REPO_ROOT / "old-files/gig-invoicing.xlsx"
OUTPUT_SQL = REPO_ROOT / "db/seeds/legacy_gigs.sql"
DUP_REPORT = REPO_ROOT / "db/seeds/legacy_gigs_dup_candidates.txt"

# ---------------------------------------------------------------------------
# Channel normalisation
# All booking platforms that route through the shared email inbox map to
# their own ENUM value for analytics; only buukkaa_bandi has a separate
# operational flow (own booking system, commission invoicing, no customer
# email exposed pre-booking).
# ---------------------------------------------------------------------------
_CHANNEL_MAP = {
    "saturday.band":          "saturday_band",
    "saturday_band":          "saturday_band",
    "buukkaa-bandi.fi":       "buukkaa_bandi",
    "www.buukkaa-bandi.fi":   "buukkaa_bandi",
    "buukkaa_bandi":          "buukkaa_bandi",
    "venuu.fi":               "venuu",
    "www.venuu.fi":           "venuu",
    "haamusiikki.com":        "haamusiikki",
    "www.haamusiikki.com":    "haamusiikki",
    "voodoolive.fi":          "voodoolive",
    "www.voodoolive.fi":      "voodoolive",
    "ohjelmanaiset.fi":       "ohjelmanaiset",
    "www.ohjelmanaiset.fi":   "ohjelmanaiset",
    "palkkaamuusikko.fi":     "palkkaamuusikko",
    "www.palkkaamuusikko.fi": "palkkaamuusikko",
    "facebook.com":           "facebook",
    "www.facebook.com":       "facebook",
    "web.whatsapp.com":       "whatsapp",
}

# ---------------------------------------------------------------------------
# Status normalisation  (patterns matched against lowercased raw string)
# "Ohi" = Finnish for "past/over" → the gig was performed.
# "Raincheckattu" = tentative/uncertain → treat as quoted (awaiting reply).
# First matching pattern wins.
# ---------------------------------------------------------------------------
_STATUS_PATTERNS = [
    (r"\bohi\b",               "delivered"),
    (r"\bvarma\b",             "confirmed"),
    (r"peruuntunut",           "cancelled"),
    (r"tarjous hylätty",       "declined"),
    (r"kieltäytynyt",          "declined"),
    (r"päivä mennyt",          "declined"),
    (r"ei halua",              "declined"),
    (r"ei vielä vastannut",    "quoted"),
    (r"raincheckattu",         "quoted"),
]

# Tokens whose presence in a name suggests a legal entity (→ customer type = company)
_COMPANY_TOKENS = frozenset({
    "oy", "ab", "llc", "ry", "yhdistys", "oyj", "ky",
    "osuuskunta", "solutions", "enterprises", "group",
})

# Gig-invoicing column indices (0-based, data rows start at row 3 in Excel)
_INV_COL_DATE    = 0   # A: KEIKKAPÄIVÄ
_INV_COL_CUST    = 2   # C: KEIKKA (customer/gig name)
_INV_COL_FEE     = 4   # E: KEIKKAPALKKIO (gross fee incl. VAT, in EUR)

# Gigs-YYYY column indices
_GIG_COL_DATE    = 0   # A: KEIKKAPÄIVÄ
_GIG_COL_INQUIRY = 2   # C: YHTEYDENOTTO
_GIG_COL_PLACE   = 4   # E: PAIKKA
_GIG_COL_CUST    = 6   # G: ASIAKAS
_GIG_COL_CHANNEL = 11  # L: KANAVA
_GIG_COL_STATUS  = 14  # O: STATUS

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class RawRecord:
    source_file:    str
    gig_date:       Optional[date]
    inquiry_date:   Optional[date]
    location:       Optional[str]
    raw_customer:   str
    channel_raw:    Optional[str]
    status_raw:     Optional[str]
    fee_eur:        Optional[float] = None   # populated from gig-invoicing only
    from_invoicing: bool = False


@dataclass
class Venue:
    id:   int
    city: str


@dataclass
class Customer:
    id:            int
    name:          str
    customer_type: str   # 'person' | 'company'


@dataclass
class Contact:
    id:         int
    first_name: str
    last_name:  str


@dataclass
class GigRow:
    id:                  int
    customer_id:         int
    contact_id:          Optional[int]
    venue_id:            Optional[int]
    gig_date:            date
    inquiry_date:        Optional[date]
    status:              str
    channel:             str
    customer_type:       str    # 'wedding' | 'company' | 'other'
    quoted_price_cents:  Optional[int]
    notes:               str


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_date(val) -> Optional[date]:
    """Coerce various date representations to date, or None."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if not isinstance(val, str):
        return None
    val = val.strip()
    # Reject values with alphabetic or placeholder characters (e.g. 'XX.01.2022', 'kesä/26')
    if re.search(r"[A-Za-zÄÖÅäöå]", val) or "X" in val.upper():
        return None
    # Finnish DD.MM.YYYY
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", val)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    # ISO YYYY-MM-DD (with optional trailing time component)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", val)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _map_channel(raw: Optional[str]) -> str:
    if not raw:
        return "mail"
    raw = str(raw).strip()
    lower = raw.lower()
    # Phone numbers: starts with '+', '(+', or is a digit-heavy string
    if re.match(r"^[\+\(]?\+?\d[\d\s\-\(\)\.]{5,}$", raw.replace(" ", "")):
        return "phone"
    # Email addresses → mail
    if "@" in raw:
        return "mail"
    # Domain lookup (substring match so 'saturday.band' matches 'saturday.band')
    for key, val in _CHANNEL_MAP.items():
        if key in lower:
            return val
    return "mail"


def _map_status(raw: Optional[str], from_invoicing: bool = False) -> str:
    if from_invoicing:
        return "delivered"
    # All historical records were written only after a quote was sent
    # (no pre-quote logging in the legacy workflow), so the correct fallback
    # for empty or unrecognised status strings is 'quoted', not 'inquiry'.
    # 'inquiry' (pre-quote) becomes relevant once automated intake is implemented.
    if not raw:
        return "quoted"
    lower = str(raw).lower()
    for pattern, status in _STATUS_PATTERNS:
        if re.search(pattern, lower):
            return status
    return "quoted"


def _normalise_name(name: str) -> str:
    """Lowercase, strip accents, remove punctuation and legal entity tokens."""
    n = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    n = n.lower()
    n = re.sub(r"[^\w\s]", "", n)
    for tok in _COMPANY_TOKENS:
        n = re.sub(rf"\b{re.escape(tok)}\b", "", n)
    return " ".join(n.split())


def _infer_customer_type(name: str) -> str:
    """'company' if the name contains a recognisable legal entity token, else 'person'."""
    lower = name.lower()
    for tok in _COMPANY_TOKENS:
        if re.search(rf"\b{re.escape(tok)}\b", lower):
            return "company"
    return "person"


def _parse_customer_contact(raw: str):
    """
    Return (customer_name, first_name, last_name, customer_entity_type).

    Handles three patterns:
      "Company Oy (Maija Haapakoski)"  → customer=company, contact extracted from parens
      "Mikael Lindqvist"               → customer=person, contact = full name split
      "[EI TIEDOSSA] (Jenni Jäntti)"   → entity unknown, contact name used as customer
    """
    if not raw:
        return None, None, None, "person"
    raw = raw.strip()

    # Pattern: "Entity (Person)" — must have non-empty content in both parts
    m = re.match(r"^(.+?)\s*\((.+?)\)\s*$", raw)
    if m:
        entity = m.group(1).strip()
        person = m.group(2).strip()

        # Nullify known-unknown placeholders
        if re.search(r"ei tiedossa|unknown|\[.*\]", entity.lower()):
            entity = ""
        if re.search(r"ei tiedossa|unknown", person.lower()):
            person = ""

        ctype = _infer_customer_type(entity) if entity else "person"
        # If entity is unknown, promote person to customer name
        customer_name = entity if entity else (person or None)
        if person:
            parts = person.split()
            first = parts[0]
            last  = " ".join(parts[1:]) if len(parts) > 1 else ""
        else:
            first = last = ""
        return customer_name, first, last, ctype

    # Bare "[EI TIEDOSSA]" with no parens → no usable name
    if re.match(r"^\[?EI TIEDOSSA\]?$", raw.upper().strip()):
        return None, None, None, "person"

    # Plain name (most common case)
    ctype = _infer_customer_type(raw)
    parts = raw.split()
    first = parts[0] if parts else raw
    last  = " ".join(parts[1:]) if len(parts) > 1 else ""
    return raw, first, last, ctype


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _load_gigs_xlsx(path: str) -> list[RawRecord]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= _GIG_COL_CUST or row[_GIG_COL_DATE] is None:
            continue
        gig_date = _parse_date(row[_GIG_COL_DATE])
        if gig_date is None:
            continue
        records.append(RawRecord(
            source_file  = os.path.basename(path),
            gig_date     = gig_date,
            inquiry_date = _parse_date(row[_GIG_COL_INQUIRY]) if len(row) > _GIG_COL_INQUIRY else None,
            location     = str(row[_GIG_COL_PLACE]).strip() if len(row) > _GIG_COL_PLACE and row[_GIG_COL_PLACE] else None,
            raw_customer = str(row[_GIG_COL_CUST]).strip() if row[_GIG_COL_CUST] else "",
            channel_raw  = str(row[_GIG_COL_CHANNEL]).strip() if len(row) > _GIG_COL_CHANNEL and row[_GIG_COL_CHANNEL] else None,
            status_raw   = str(row[_GIG_COL_STATUS]).strip() if len(row) > _GIG_COL_STATUS and row[_GIG_COL_STATUS] else None,
        ))
    return records


def _load_invoicing_xlsx(path: str) -> list[RawRecord]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    # Row 2 is the header; data starts at row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[_INV_COL_DATE] is None:
            continue
        gig_date = _parse_date(row[_INV_COL_DATE])
        if gig_date is None:
            continue
        fee = None
        if len(row) > _INV_COL_FEE and row[_INV_COL_FEE] is not None:
            try:
                fee = float(row[_INV_COL_FEE])
            except (TypeError, ValueError):
                pass
        records.append(RawRecord(
            source_file    = "gig-invoicing.xlsx",
            gig_date       = gig_date,
            inquiry_date   = None,
            location       = None,
            raw_customer   = str(row[_INV_COL_CUST]).strip() if row[_INV_COL_CUST] else "",
            channel_raw    = None,
            status_raw     = None,
            fee_eur        = fee,
            from_invoicing = True,
        ))
    return records


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def _is_gig_match(a: RawRecord, b: RawRecord) -> bool:
    """True if two records are likely the same real-world gig.

    Criteria: date within ±3 days AND normalised name similarity ≥ 0.75.
    The ±3-day window accounts for occasional date discrepancies between the
    invoicing spreadsheet and the booking tracker.
    """
    if a.gig_date is None or b.gig_date is None:
        return False
    date_diff = abs((a.gig_date - b.gig_date).days)
    sim = SequenceMatcher(
        None,
        _normalise_name(a.raw_customer),
        _normalise_name(b.raw_customer),
    ).ratio()
    return date_diff <= 3 and sim >= 0.75


def _find_dup_name_candidates(records: list[RawRecord]) -> list[tuple[RawRecord, RawRecord]]:
    """Return pairs whose names are similar but not identical (potential duplicates)."""
    pairs = []
    seen: set[tuple[int, int]] = set()
    for i, a in enumerate(records):
        na = _normalise_name(a.raw_customer)
        if not na:
            continue
        for j, b in enumerate(records):
            if j <= i:
                continue
            key = (i, j)
            if key in seen:
                continue
            nb = _normalise_name(b.raw_customer)
            if na == nb or not nb:
                continue
            if SequenceMatcher(None, na, nb).ratio() >= 0.82:
                pairs.append((a, b))
                seen.add(key)
    return pairs


# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------

def _sq(val) -> str:
    """Wrap value in single quotes for SQL, escaping internal quotes. Returns NULL for None."""
    if val is None:
        return "NULL"
    return "'" + str(val).replace("\\", "\\\\").replace("'", "\\'") + "'"


def _sd(val: Optional[date]) -> str:
    return f"'{val.isoformat()}'" if val else "NULL"


def _si(val) -> str:
    return str(int(val)) if val is not None else "NULL"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Extract legacy gig data to SQL seed.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print generated SQL to stdout instead of writing files.")
    parser.add_argument("--stats", action="store_true",
                        help="Print entity counts and exit without writing anything.")
    args = parser.parse_args()

    # ── 1. Load sources ──────────────────────────────────────────────────────
    gigs_yyyy: list[RawRecord] = []
    for pattern in GIGS_GLOBS:
        for f in sorted(glob.glob(pattern)):
            recs = _load_gigs_xlsx(f)
            print(f"  {len(recs):3d} records  ← {os.path.basename(f)}", file=sys.stderr)
            gigs_yyyy.extend(recs)

    invoicing_recs: list[RawRecord] = []
    if INVOICING_PATH.exists():
        invoicing_recs = _load_invoicing_xlsx(str(INVOICING_PATH))
        print(f"  {len(invoicing_recs):3d} records  ← gig-invoicing.xlsx", file=sys.stderr)

    # ── 2. Merge gig-invoicing into gigs_yyyy ────────────────────────────────
    # For each invoicing record that matches a gigs-YYYY record: propagate the
    # fee to the gigs-YYYY record (it has richer metadata).
    # For invoicing records with no match: include them as additional delivered gigs.
    unmatched_invoicing: list[RawRecord] = []
    for inv in invoicing_recs:
        matched_idx = next(
            (i for i, g in enumerate(gigs_yyyy) if _is_gig_match(inv, g)),
            None
        )
        if matched_idx is not None:
            g = gigs_yyyy[matched_idx]
            if g.fee_eur is None and inv.fee_eur is not None:
                gigs_yyyy[matched_idx] = RawRecord(
                    source_file=g.source_file, gig_date=g.gig_date,
                    inquiry_date=g.inquiry_date, location=g.location,
                    raw_customer=g.raw_customer, channel_raw=g.channel_raw,
                    status_raw=g.status_raw, fee_eur=inv.fee_eur,
                    from_invoicing=False,
                )
        else:
            unmatched_invoicing.append(inv)

    n_matched = len(invoicing_recs) - len(unmatched_invoicing)
    print(f"\n  gig-invoicing: {n_matched} matched to gigs-YYYY, "
          f"{len(unmatched_invoicing)} unmatched → added as extra delivered gigs",
          file=sys.stderr)

    all_records = gigs_yyyy + unmatched_invoicing

    # ── 3. Build entity maps ─────────────────────────────────────────────────
    # Each map key is the normalised name; value is the entity dataclass.
    venue_map:    dict[str, Venue]    = {}
    customer_map: dict[str, Customer] = {}
    contact_map:  dict[str, Contact]  = {}

    def get_or_create_venue(city: Optional[str]) -> Optional[int]:
        if not city:
            return None
        key = city.strip().lower()
        if key not in venue_map:
            venue_map[key] = Venue(id=len(venue_map) + 1, city=city.strip())
        return venue_map[key].id

    def get_or_create_customer(name: Optional[str], ctype: str) -> Optional[int]:
        if not name:
            return None
        key = _normalise_name(name)
        if not key:
            return None
        if key not in customer_map:
            customer_map[key] = Customer(id=len(customer_map) + 1,
                                         name=name, customer_type=ctype)
        return customer_map[key].id

    def get_or_create_contact(first: str, last: str) -> Optional[int]:
        full = f"{first} {last}".strip()
        if not full:
            return None
        key = _normalise_name(full)
        if not key:
            return None
        if key not in contact_map:
            contact_map[key] = Contact(id=len(contact_map) + 1,
                                       first_name=first, last_name=last)
        return contact_map[key].id

    customer_contact_pairs: set[tuple[int, int]] = set()

    # ── 4. Build gig rows ────────────────────────────────────────────────────
    gig_rows: list[GigRow] = []
    skipped = 0
    for rec in all_records:
        cust_name, first, last, ctype = _parse_customer_contact(rec.raw_customer)
        customer_id = get_or_create_customer(cust_name, ctype)
        if customer_id is None:
            # Cannot satisfy NOT NULL constraint; skip with a warning.
            print(f"  SKIP (no customer): {rec.source_file} {rec.gig_date} "
                  f"{rec.raw_customer!r}", file=sys.stderr)
            skipped += 1
            continue

        contact_id = get_or_create_contact(first, last)
        venue_id   = get_or_create_venue(rec.location)

        if contact_id is not None:
            customer_contact_pairs.add((customer_id, contact_id))

        # Derive gig customer_type from the entity type heuristic.
        # Defaults to 'wedding' for persons since that is the band's primary market.
        # Companies get 'company'; 'other' is not assigned automatically.
        gig_customer_type = "company" if ctype == "company" else "wedding"

        # Heuristic note: all statuses other than gig-invoicing records are mapped
        # from Finnish free-text strings; see _STATUS_PATTERNS for logic.
        status  = _map_status(rec.status_raw, rec.from_invoicing)
        channel = _map_channel(rec.channel_raw)

        # Treat fee ≤ 0 as absent (can result from empty formula cells in Excel)
        quoted_price_cents = (
            round(rec.fee_eur * 100) if rec.fee_eur and rec.fee_eur > 0 else None
        )

        # Use inquiry_date as created_at when available (closest proxy for when
        # the ERP record would have been created in the new system).
        notes = f"Legacy import: {rec.source_file}"
        if rec.from_invoicing:
            notes += " (no matching gigs-YYYY record)"

        gig_rows.append(GigRow(
            id                 = len(gig_rows) + 1,
            customer_id        = customer_id,
            contact_id         = contact_id,
            venue_id           = venue_id,
            gig_date           = rec.gig_date,
            inquiry_date       = rec.inquiry_date,
            status             = status,
            channel            = channel,
            customer_type      = gig_customer_type,
            quoted_price_cents = quoted_price_cents,
            notes              = notes,
        ))

    # ── 5. Stats ─────────────────────────────────────────────────────────────
    print(f"\n  Entities: {len(venue_map)} venues · {len(customer_map)} customers · "
          f"{len(contact_map)} contacts · {len(gig_rows)} gigs "
          f"({skipped} skipped)", file=sys.stderr)

    if args.stats:
        return 0

    # ── 6. Duplicate candidate report ────────────────────────────────────────
    dup_pairs = _find_dup_name_candidates(all_records)

    # ── 7. Generate SQL ───────────────────────────────────────────────────────
    now_ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    safe_ai = max(
        max((v.id for v in venue_map.values()), default=0),
        max((c.id for c in customer_map.values()), default=0),
        max((c.id for c in contact_map.values()), default=0),
        max((g.id for g in gig_rows), default=0),
        9999,
    ) + 1

    lines: list[str] = []

    lines.append(f"""\
-- ===========================================================================
-- LEGACY GIGS IMPORT
-- Generated by cli/etl/extract_gigs.py  on  {now_ts} UTC
-- Sources: old-files/info/gigs-*.xlsx (+ archive/)  +  old-files/gig-invoicing.xlsx
--
-- IDs 1–9999 are reserved for legacy import data.
-- Manually-created ERP records will start from AUTO_INCREMENT = {safe_ai}.
--
-- Apply db/migrations/001_expand_channel_enum.sql before loading this file.
-- Re-running is safe: the DELETE block clears all legacy rows first.
-- ===========================================================================

SET FOREIGN_KEY_CHECKS = 0;

-- Clear previous legacy import
DELETE FROM gigs              WHERE id BETWEEN 1 AND 9999;
DELETE FROM customer_contacts WHERE customer_id BETWEEN 1 AND 9999;
DELETE FROM contacts          WHERE id BETWEEN 1 AND 9999;
DELETE FROM customers         WHERE id BETWEEN 1 AND 9999;
DELETE FROM venues            WHERE id BETWEEN 1 AND 9999;
""")

    # venues
    lines.append("-- ---------------------------------------------------------------------------")
    lines.append("-- Venues  (city-level; name = city until enriched in ERP)")
    lines.append("-- ---------------------------------------------------------------------------")
    venue_vals = ",\n".join(
        f"  ({v.id}, {_sq(v.city)}, {_sq(v.city)}, 'FI', '{now_ts}', '{now_ts}')"
        for v in sorted(venue_map.values(), key=lambda x: x.id)
    )
    lines.append(f"INSERT INTO venues (id, name, city, country, created_at, updated_at) VALUES\n{venue_vals};")
    lines.append("")

    # customers
    lines.append("-- ---------------------------------------------------------------------------")
    lines.append("-- Customers  (type heuristic: contains legal entity token → company)")
    lines.append("-- ---------------------------------------------------------------------------")
    cust_vals = ",\n".join(
        f"  ({c.id}, {_sq(c.name)}, {_sq(c.customer_type)}, '{now_ts}', '{now_ts}')"
        for c in sorted(customer_map.values(), key=lambda x: x.id)
    )
    lines.append(f"INSERT INTO customers (id, name, type, created_at, updated_at) VALUES\n{cust_vals};")
    lines.append("")

    # contacts
    lines.append("-- ---------------------------------------------------------------------------")
    lines.append("-- Contacts  (name-only stubs; email/phone to be filled in ERP)")
    lines.append("-- ---------------------------------------------------------------------------")
    cont_vals = ",\n".join(
        f"  ({c.id}, {_sq(c.first_name)}, {_sq(c.last_name)}, '{now_ts}', '{now_ts}')"
        for c in sorted(contact_map.values(), key=lambda x: x.id)
    )
    lines.append(f"INSERT INTO contacts (id, first_name, last_name, created_at, updated_at) VALUES\n{cont_vals};")
    lines.append("")

    # customer_contacts
    lines.append("-- ---------------------------------------------------------------------------")
    lines.append("-- Customer–contact bridge")
    lines.append("-- ---------------------------------------------------------------------------")
    cc_vals = ",\n".join(
        f"  ({cid}, {ctid}, 1, '{now_ts}')"
        for cid, ctid in sorted(customer_contact_pairs)
    )
    lines.append(f"INSERT INTO customer_contacts (customer_id, contact_id, is_primary, created_at) VALUES\n{cc_vals};")
    lines.append("")

    # gigs
    lines.append("-- ---------------------------------------------------------------------------")
    lines.append("-- Gigs")
    lines.append("-- customer_type heuristic defaults to 'wedding' for individual persons.")
    lines.append("-- quoted_price_cents populated only where gig-invoicing.xlsx data is available.")
    lines.append("-- created_at = inquiry date when known, otherwise script run timestamp.")
    lines.append("-- ---------------------------------------------------------------------------")
    gig_vals = ",\n".join(
        f"  ({g.id}, {_si(g.customer_id)}, {_si(g.contact_id)}, {_si(g.venue_id)}, "
        f"{_sd(g.gig_date)}, {_sq(g.status)}, {_sq(g.channel)}, "
        f"{_sq(g.customer_type)}, {_si(g.quoted_price_cents)}, {_sq(g.notes)}, "
        f"{_sq((g.inquiry_date or g.gig_date).strftime('%Y-%m-%d %H:%M:%S') if (g.inquiry_date or g.gig_date) else now_ts)}, "
        f"'{now_ts}')"
        for g in gig_rows
    )
    lines.append(
        "INSERT INTO gigs\n"
        "  (id, customer_id, contact_id, venue_id, gig_date, status, channel,\n"
        "   customer_type, quoted_price_cents, notes, created_at, updated_at)\n"
        f"VALUES\n{gig_vals};"
    )
    lines.append("")

    # reset AUTO_INCREMENT
    lines.append(f"""\
SET FOREIGN_KEY_CHECKS = 1;

-- Ensure ERP-created records start above the legacy ID range
ALTER TABLE venues         AUTO_INCREMENT = {safe_ai};
ALTER TABLE customers      AUTO_INCREMENT = {safe_ai};
ALTER TABLE contacts       AUTO_INCREMENT = {safe_ai};
ALTER TABLE gigs           AUTO_INCREMENT = {safe_ai};
""")

    sql = "\n".join(lines)

    # ── 8. Write output ───────────────────────────────────────────────────────
    if args.dry_run:
        print(sql)
    else:
        OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)
        OUTPUT_SQL.write_text(sql, encoding="utf-8")
        print(f"\n  Wrote {OUTPUT_SQL.relative_to(REPO_ROOT)}", file=sys.stderr)

    # Duplicate report
    if dup_pairs:
        report_lines = [
            "# Near-duplicate customer name candidates",
            "# These pairs have normalised name similarity ≥ 0.82.",
            "# Review in the ERP and merge manually if they represent the same person.",
            f"# Generated: {now_ts} UTC",
            f"# Total pairs: {len(dup_pairs)}",
            "",
        ]
        for a, b in dup_pairs:
            report_lines.append(
                f"  {a.raw_customer!r:<45s}  ↔  {b.raw_customer!r}"
                f"  [{a.source_file} / {b.source_file}]"
            )
        dup_report_text = "\n".join(report_lines)

        if args.dry_run:
            print("\n--- Duplicate candidates ---", file=sys.stderr)
            print(dup_report_text, file=sys.stderr)
        else:
            DUP_REPORT.write_text(dup_report_text, encoding="utf-8")
            print(f"  Wrote {DUP_REPORT.relative_to(REPO_ROOT)} "
                  f"({len(dup_pairs)} pairs)", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
