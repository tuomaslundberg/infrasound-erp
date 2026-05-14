#!/usr/bin/env python3
"""
cli/etl/extract_invoicing.py

Build gig_personnel rows and update quoted_price_cents from gig-invoicing.xlsx.

Strategy
--------
For each row in gig-invoicing.xlsx:
  1. Fuzzy-match to a DB gig by (gig_date, customer_name).
  2. If keikkapalkkio > 0: emit UPDATE gigs SET quoted_price_cents.
  3. Emit DELETE + INSERT for gig_personnel:
     - Partners (Tuomas, Toni/Valtteri, Joni, Lauri) present by default.
     - External musicians (Mikael, Emil, Alina, Mortti, Samuel, Leevi) if fee > 0.
     - MUUT column: hardcoded date→musician mapping (confirmed via calendar log).
     - Partner exceptions and fee-in-KULUT substitutes: hardcoded per spec.

Source
------
  old-files/gig-invoicing.xlsx  (64 gigs, 2020-02 → 2025-12)

Prerequisites
-------------
  - migrations 006 + 007 applied (fee_cents nullable; role ENUM updated)
  - db/seeds/musicians.sql loaded (all 19 roster members seeded)
  - Gigs loaded (make import-legacy-gigs)

Output
------
  db/seeds/legacy_invoicing.sql
  db/seeds/legacy_invoicing_unmatched.txt

Usage
-----
  python cli/etl/extract_invoicing.py
  python cli/etl/extract_invoicing.py --dry-run
  python cli/etl/extract_invoicing.py --stats
"""

from __future__ import annotations

import argparse
import difflib
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Optional

REPO_ROOT      = Path(__file__).resolve().parent.parent.parent
SOURCE_XLSX    = REPO_ROOT / "old-files" / "gig-invoicing.xlsx"
OUTPUT_SQL     = REPO_ROOT / "db" / "seeds" / "legacy_invoicing.sql"
UNMATCHED_FILE = REPO_ROOT / "db" / "seeds" / "legacy_invoicing_unmatched.txt"

FUZZY_THRESHOLD = 0.75

# ---------------------------------------------------------------------------
# Column indices (1-based, openpyxl)
# Every data column occupies an odd column; even columns are blank/merged.
# ---------------------------------------------------------------------------
COL_DATE     = 1
COL_CUST     = 3
COL_GROSS    = 5
# COL_NET    = 7   # deferred
COL_ENGINEER = 9   # MIKSAAJAN PALKKIO (Toni or Valtteri)
# COL_KULUT  = 11  # deferred to Phase 7
COL_MIKAEL   = 15
COL_EMIL     = 17
COL_ALINA    = 19
COL_MORTTI   = 21
COL_SAMUEL   = 23
COL_LEEVI    = 25
COL_MUUT     = 27

# External musician columns: (col, username, role)
EXTERNAL_COLS = [
    (COL_MIKAEL, "mikael.lehto",       "vocals"),
    (COL_EMIL,   "emil.lamminmaki",    "bass"),
    (COL_ALINA,  "alina.kangas",       "vocals"),
    (COL_MORTTI, "mortti.markkanen",   "bass"),
    (COL_SAMUEL, "samuel.johansson",   "bass"),
    (COL_LEEVI,  "leevi.kahkonen",     "bass"),
]

# ---------------------------------------------------------------------------
# Exception tables (all confirmed against spec / calendar log)
# ---------------------------------------------------------------------------

# Partners absent from specific gigs.
# Toni absence on Valtteri gigs is handled via VALTTERI_DATES, not here.
PARTNER_ABSENT: dict[str, set[str]] = {
    "2022-07-15": {"toni.puttonen", "joni.virtanen"},   # trio gig
    "2022-07-16": {"joni.virtanen"},                     # Erkki Sippel sub
    "2022-07-23": {"toni.puttonen"},                     # Leevi sub
    "2022-07-30": {"toni.puttonen"},                     # Leevi sub
    "2025-06-14": {"tuomas.lundberg"},                   # Eetu Hämäläinen sub; Toni via VALTTERI
    "2025-07-12": {"tuomas.lundberg"},                   # Juho Peuraniemi sub via MUUT
}

# Dates where Valtteri Alanen fills the sound_engineering slot (Toni absent).
# Fee for Valtteri comes from COL_ENGINEER, same as Toni's slot.
VALTTERI_DATES: set[str] = {
    "2025-06-14", "2025-06-28", "2025-07-05",
    "2025-07-26", "2025-09-13", "2025-11-14",
}

# Substitutes whose fees are embedded in KULUT (not in a dedicated column).
# fee_cents = None → stored as NULL (KULUT deferred).
SUBSTITUTES: dict[str, list[tuple[str, str, Optional[int]]]] = {
    "2022-07-16": [("erkki.sippel",     "drums",             None)],
    "2022-07-23": [("leevi.kahkonen",   "sound_engineering", 12097)],  # 120.97 € net
    "2022-07-30": [("leevi.kahkonen",   "sound_engineering", 16935)],  # 169.35 € net
    "2023-07-28": [("erkki.sippel",     "bass",              None)],
    "2023-07-29": [("antti.saari",      "bass",              None)],
    "2025-06-14": [("eetu.hamalainen",  "keyboards",         None)],   # fee in KULUT
}

# MUUT column (COL_MUUT) → musician identity, confirmed via calendar.
MUUT_MAP: dict[str, tuple[str, str]] = {
    "2024-08-03": ("lassi.kriikkula",   "bass"),
    "2024-08-24": ("maxwell.mbare",     "bass"),
    "2024-09-12": ("iris.toivonen",     "vocals"),
    "2024-09-21": ("maxwell.mbare",     "bass"),
    "2024-10-12": ("maxwell.mbare",     "bass"),
    "2025-06-14": ("maxwell.mbare",     "bass"),
    "2025-07-12": ("juho.peuraniemi",   "keyboards"),
    "2025-07-26": ("maxwell.mbare",     "bass"),
    "2025-08-16": ("arttu.luonsinen",   "bass"),
    "2025-08-23": ("maxwell.mbare",     "bass"),
    "2025-11-14": ("maxwell.mbare",     "bass"),
}

# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return " ".join(s.split())


def _parse_date(raw) -> Optional[str]:
    """Return ISO date string (YYYY-MM-DD) or None."""
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    # DD.MM.YYYY
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return None


def _to_cents(val) -> Optional[int]:
    if val is None:
        return None
    try:
        f = float(val)
        if f == 0.0:
            return None
        return round(f * 100)
    except (TypeError, ValueError):
        return None

# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------

def _sq(val: Optional[str]) -> str:
    if val is None:
        return "NULL"
    return "'" + str(val).replace("\\", "\\\\").replace("'", "\\'") + "'"

# ---------------------------------------------------------------------------
# .env / DB helpers (mirrors other ETL scripts)
# ---------------------------------------------------------------------------

def _load_dot_env(path: Optional[Path] = None) -> dict[str, str]:
    if path is None:
        path = REPO_ROOT / ".env"
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def _get_env() -> dict[str, str]:
    env = _load_dot_env()
    dev = REPO_ROOT / ".env.dev"
    if dev.exists():
        env.update(_load_dot_env(dev))
    return env


def _try_connect_db(env: dict[str, str]):
    try:
        import pymysql  # type: ignore
    except ImportError:
        return None
    host     = env.get("ETL_DB_HOST") or env.get("DB_HOST", "127.0.0.1")
    port     = int(env.get("ETL_DB_PORT") or env.get("DB_PORT", "3306"))
    user     = env.get("ETL_DB_USER")     or env.get("MYSQL_USER")
    password = env.get("ETL_DB_PASSWORD") or env.get("MYSQL_PASSWORD")
    dbname   = env.get("MYSQL_DATABASE")
    if not all([user, password, dbname]):
        return None
    try:
        return pymysql.connect(
            host=host, port=port, user=user, password=password,
            database=dbname, charset="utf8mb4", connect_timeout=5,
        )
    except Exception as e:
        print(f"  DB connect failed: {e}", file=sys.stderr)
        return None

# ---------------------------------------------------------------------------
# DB loaders
# ---------------------------------------------------------------------------

def load_gigs(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT g.id, g.gig_date, c.name AS customer_name "
            "FROM gigs g "
            "LEFT JOIN customers c ON c.id = g.customer_id "
            "WHERE g.id BETWEEN 1 AND 9999 AND g.deleted_at IS NULL"
        )
        return [
            {"id": r[0], "gig_date": r[1], "customer_name": r[2] or ""}
            for r in cur.fetchall()
        ]


def load_users(conn) -> dict[str, int]:
    """Return {username: id}."""
    with conn.cursor() as cur:
        cur.execute("SELECT id, username FROM users WHERE deleted_at IS NULL")
        return {r[1]: r[0] for r in cur.fetchall()}

# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def _match_gig(
    inv_date: str,
    inv_cust: str,
    gigs: list[dict],
) -> Optional[dict]:
    """Fuzzy-match an invoicing row to a DB gig by date + customer name."""
    inv_date_obj = datetime.strptime(inv_date, "%Y-%m-%d").date()
    candidates = [
        g for g in gigs
        if g["gig_date"] is not None and abs((g["gig_date"] - inv_date_obj).days) <= 3
    ]
    if not candidates:
        return None
    inv_key = _norm(inv_cust)
    best = None
    best_score = 0.0
    for g in candidates:
        score = difflib.SequenceMatcher(None, inv_key, _norm(g["customer_name"])).ratio()
        if score > best_score:
            best_score = score
            best = g
    if best and best_score >= FUZZY_THRESHOLD:
        return best
    return None

# ---------------------------------------------------------------------------
# Personnel builder
# ---------------------------------------------------------------------------

def build_personnel(
    date_str: str,
    ws_row,
    ws,
    user_map: dict[str, int],
    warnings: list[str],
) -> list[dict]:
    """
    Return list of {username, user_id, role, fee_cents} for one invoicing row.
    """
    absent: set[str] = PARTNER_ABSENT.get(date_str, set()).copy()

    # Sound engineering slot
    if date_str in VALTTERI_DATES:
        engineer_username = "valtteri.alanen"
        absent.add("toni.puttonen")
    elif "toni.puttonen" in absent:
        engineer_username = None  # substitute via SUBSTITUTES; col 9 fee is 0
    else:
        engineer_username = "toni.puttonen"

    rows: list[dict] = []

    def _add(username: str, role: str, fee_cents: Optional[int]):
        uid = user_map.get(username)
        if uid is None:
            warnings.append(f"  WARN: user not found: {username}")
            return
        rows.append({"username": username, "user_id": uid, "role": role, "fee_cents": fee_cents})

    # Partners
    for (username, role) in [
        ("tuomas.lundberg", "keyboards"),
        ("joni.virtanen",   "drums"),
        ("lauri.lehtinen",  "guitar"),
    ]:
        if username not in absent:
            _add(username, role, None)  # partner fees not individually tracked

    # Sound engineer
    if engineer_username:
        eng_fee = _to_cents(ws.cell(ws_row, COL_ENGINEER).value)
        _add(engineer_username, "sound_engineering", eng_fee)

    # External musicians by fee column
    for (col, username, role) in EXTERNAL_COLS:
        fee = _to_cents(ws.cell(ws_row, col).value)
        if fee is not None:
            _add(username, role, fee)

    # MUUT column
    muut_fee = _to_cents(ws.cell(ws_row, COL_MUUT).value)
    if muut_fee is not None:
        if date_str in MUUT_MAP:
            muut_username, muut_role = MUUT_MAP[date_str]
            _add(muut_username, muut_role, muut_fee)
        else:
            warnings.append(
                f"  WARN: MUUT fee {muut_fee} on {date_str} has no mapping — skipped"
            )

    # Hardcoded substitutes (fee in KULUT)
    for (username, role, fee_cents) in SUBSTITUTES.get(date_str, []):
        _add(username, role, fee_cents)

    return rows

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build gig_personnel + quoted_price_cents from gig-invoicing.xlsx."
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Print SQL to stdout instead of writing file.")
    parser.add_argument("--stats", action="store_true",
                        help="Report match counts only; write no files.")
    args = parser.parse_args()

    if not SOURCE_XLSX.exists():
        print(f"  ERROR: source file not found: {SOURCE_XLSX}", file=sys.stderr)
        return 1

    env = _get_env()
    conn = _try_connect_db(env)
    if conn is None:
        print("  ERROR: Cannot connect to DB.", file=sys.stderr)
        return 1

    gigs     = load_gigs(conn)
    user_map = load_users(conn)
    conn.close()
    print(f"  DB: {len(gigs)} legacy gigs, {len(user_map)} users", file=sys.stderr)

    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("  ERROR: openpyxl not installed.  Run: pip install openpyxl", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.active

    matched_rows:   list[dict] = []
    unmatched_rows: list[str]  = []
    warnings:       list[str]  = []

    for row in range(3, ws.max_row + 1):
        raw_date = ws.cell(row, COL_DATE).value
        raw_cust = ws.cell(row, COL_CUST).value
        if raw_date is None and raw_cust is None:
            continue

        date_str = _parse_date(raw_date)
        cust_str = str(raw_cust).strip() if raw_cust else ""
        if not date_str or not cust_str:
            unmatched_rows.append(f"PARSE_FAIL\t{raw_date}\t{raw_cust}")
            continue

        gig = _match_gig(date_str, cust_str, gigs)
        if gig is None:
            unmatched_rows.append(f"NO_MATCH\t{date_str}\t{cust_str}")
            continue

        gross_cents = _to_cents(ws.cell(row, COL_GROSS).value)
        personnel   = build_personnel(date_str, row, ws, user_map, warnings)

        matched_rows.append({
            "gig_id":      gig["id"],
            "date_str":    date_str,
            "cust_str":    cust_str,
            "gross_cents": gross_cents,
            "personnel":   personnel,
        })

    print(f"  Matched: {len(matched_rows)}", file=sys.stderr)
    print(f"  Unmatched: {len(unmatched_rows)}", file=sys.stderr)
    for w in warnings:
        print(w, file=sys.stderr)

    if args.stats:
        return 0

    # Build SQL
    now_ts = __import__("datetime").datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "-- ===========================================================================",
        "-- INVOICING ETL SEED",
        f"-- Generated by cli/etl/extract_invoicing.py  on  {now_ts} UTC",
        "--",
        "-- Updates quoted_price_cents on gigs; populates gig_personnel.",
        "-- Safe to re-run: DELETE + INSERT pattern per gig.",
        "-- ===========================================================================",
        "",
        "SET NAMES utf8mb4;",
        "",
    ]

    for m in matched_rows:
        gig_id = m["gig_id"]
        lines.append(f"-- {m['date_str']}  {m['cust_str']}")

        if m["gross_cents"] is not None:
            lines.append(
                f"UPDATE gigs SET quoted_price_cents = {m['gross_cents']} "
                f"WHERE id = {gig_id};"
            )

        lines.append(f"DELETE FROM gig_personnel WHERE gig_id = {gig_id};")
        if m["personnel"]:
            vals = []
            for p in m["personnel"]:
                fee = str(p["fee_cents"]) if p["fee_cents"] is not None else "NULL"
                vals.append(f"({gig_id}, {p['user_id']}, '{p['role']}', {fee}, NULL)")
            lines.append(
                f"INSERT INTO gig_personnel (gig_id, user_id, role, fee_cents, confirmed_at) VALUES"
            )
            lines.append(",\n  ".join(f"  {v}" for v in vals) + ";")

        lines.append("")

    sql = "\n".join(lines)

    if args.dry_run:
        print(sql)
    else:
        OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)
        OUTPUT_SQL.write_text(sql, encoding="utf-8")
        print(f"  Wrote {OUTPUT_SQL.relative_to(REPO_ROOT)}", file=sys.stderr)

    if unmatched_rows and not args.dry_run:
        UNMATCHED_FILE.write_text("\n".join(unmatched_rows) + "\n", encoding="utf-8")
        print(f"  Unmatched: {UNMATCHED_FILE.relative_to(REPO_ROOT)}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
